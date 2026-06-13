from __future__ import annotations

import datetime as dt
import hashlib
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_XLSX = Path(r"C:\Users\micha\Downloads\BD_VERSAO_VEGANA (1).xlsx")
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "supabase" / "seed" / "production_import.local.sql"
EXCEL_EPOCH = dt.datetime(1899, 12, 30)


def fix_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    try:
        value = value.encode("latin1").decode("utf-8")
    except UnicodeError:
        pass
    return value.strip()


def as_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def as_decimal(value: object, default: str | None = "0") -> str | None:
    if value in (None, ""):
        return default
    try:
        return str(Decimal(str(value)).quantize(Decimal("0.001")).normalize())
    except (InvalidOperation, ValueError):
        return default


def as_phone(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d+(\.\d+)?E\+?\d+", text, flags=re.I):
        text = str(int(float(text)))
    digits = re.sub(r"\D", "", text)
    return digits or text


def excel_datetime(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.replace(tzinfo=None).isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    if serial <= 0 or serial > 100000:
        return None
    return (EXCEL_EPOCH + dt.timedelta(days=serial)).isoformat()


def sql(value: object) -> str:
    value = fix_text(value)
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value).strip()
    if text == "":
        return "null"
    return "'" + text.replace("'", "''") + "'"


def status_pedido(value: object) -> str:
    text = str(value or "").strip().lower()
    if text == "finalizado":
        return "finalizado"
    if text == "cancelado":
        return "cancelado"
    if text == "rascunho":
        return "rascunho"
    return "pendente"


def normalize(value: object) -> str:
    import unicodedata

    text = str(fix_text(value) or "").lower()
    return "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")


def estimate_yield(product_name: object, category: object, weight: object) -> int:
    name = normalize(product_name)
    cat = normalize(category)
    grams = Decimal(as_decimal(weight, "0") or "0")

    if "bebida" in cat or re.search(r"\b(agua|coca|suco|cha|refrigerante)\b", name):
        return 1
    if re.search(r"\b(bolo|torta|travessa|bombom|sobremesa|pave|mousse)\b", name):
        return 12
    if re.search(r"\b(baguete|caldo|sopa|canjica|feijoada|baiao|moqueca|macarrao|brasileirinho|bife|almond|almondega)\b", name):
        return 20
    if grams and grams <= 100:
        return 1
    if grams and grams <= 500:
        return 2
    return 20 if "prato" in cat else 10


def classify_order(resource_name: object) -> int:
    name = normalize(resource_name)
    if re.search(r"(agua|caldo|leite|molho|extrato|vinagre)", name):
        return 10
    if re.search(r"(arroz|feijao|lentilha|grao|quinoa|macarrao|massa|farinha|aveia|proteina|soja|tofu|seitan|batata|mandioca|inhame|banana)", name):
        return 20
    if re.search(r"(cenoura|tomate|abobora|beterraba|milho|ervilha|brocolis|couve|repolho|pimentao|abobrinha|berinjela|palmito|cogumelo)", name):
        return 30
    if re.search(r"(oleo|azeite|cebola|alho|alho poro|gengibre|castanha|amendoim|gergelim)", name):
        return 40
    if re.search(r"(sal|pimenta|paprica|cominho|curry|acafrao|oregano|louro|noz|canela|tempero|ervas)", name):
        return 50
    if re.search(r"(limao|salsa|coentro|cebolinha|cheiro verde|hortela|manjericao|folha|creme|coco)", name):
        return 60
    return 70


def stable_hash(values: list[object]) -> str:
    text = "|".join(str(fix_text(value) or "") for value in values)
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def rows(ws):
    iterator = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration:
        return []
    headers = [str(item).strip() if item is not None else "" for item in raw_headers]
    output = []
    for excel_row_number, raw in enumerate(iterator, start=2):
        record = {headers[index]: raw[index] if index < len(raw) else None for index in range(len(headers)) if headers[index]}
        if any(value not in (None, "") for value in record.values()):
            record["_linha_excel"] = excel_row_number
            output.append(record)
    return output


def insert(table: str, columns: list[str], values: list[object], conflict: str | None = None) -> str:
    rendered_values = ", ".join(sql(value) for value in values)
    rendered_columns = ", ".join(columns)
    statement = f"insert into public.{table} ({rendered_columns}) values ({rendered_values})"
    if conflict:
        update_columns = [column for column in columns if column != conflict]
        updates = ", ".join(f"{column} = excluded.{column}" for column in update_columns)
        statement += f" on conflict ({conflict}) do update set {updates}"
    return statement + ";"


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    statements = [
        "-- Arquivo local gerado a partir da planilha de producao. Nao versionar.",
        "begin;",
    ]
    counts: dict[str, int] = {}

    def add(statement: str, table: str) -> None:
        statements.append(statement)
        counts[table] = counts.get(table, 0) + 1

    for row in rows(wb["Categorias"]):
        cod_categoria = as_int(row.get("COD_CATEGORIA"))
        if cod_categoria is not None:
            add(insert("categorias", ["cod_categoria", "categoria"], [cod_categoria, row.get("CATEGORIA")], "cod_categoria"), "categorias")

    for row in rows(wb["Tipo Medida"]):
        tipo_medida = row.get("TIPO_MEDIDA")
        if tipo_medida:
            add(insert("tipos_medida", ["tipo_medida"], [tipo_medida], "tipo_medida"), "tipos_medida")

    for row in rows(wb["Regiões"]):
        id_regiao = as_int(row.get("ID_REGIAO"))
        if id_regiao is not None:
            add(insert("regioes", ["id_regiao", "regiao", "taxa"], [id_regiao, row.get("REGIAO"), as_decimal(row.get("TAXA"))], "id_regiao"), "regioes")

    for row in rows(wb["Forma Pagamento"]):
        id_forma_pagamento = as_int(row.get("ID_FORMA_PAGAMENTO"))
        if id_forma_pagamento is not None:
            add(insert("formas_pagamento", ["id_forma_pagamento", "forma_pagamento", "bandeira", "prazo_recebimento", "taxa"], [id_forma_pagamento, row.get("FORMA_PAGAMENTO"), row.get("BANDEIRA"), row.get("PRAZO_RECEBIMENTO"), as_decimal(row.get("TAXA"))], "id_forma_pagamento"), "formas_pagamento")

    for row in rows(wb["Fornecedores"]):
        id_fornecedor = as_int(row.get("ID_FORNECEDOR"))
        if id_fornecedor is not None:
            add(insert("fornecedores", ["id_fornecedor", "nome_fornecedor", "telefone_fornecedor", "email_fornecedor", "observacao", "data_inclusao"], [id_fornecedor, row.get("NOME_FORNECEDOR"), as_phone(row.get("TELEFONE_FORNECEDOR")), row.get("EMAIL_FORNECEDOR"), row.get("OBSERVACAO"), excel_datetime(row.get("DATA_INCLUSAO"))], "id_fornecedor"), "fornecedores")

    for row in rows(wb["Recursos"]):
        id_recurso = as_int(row.get("ID_RECURSO"))
        if id_recurso is not None:
            add(insert("recursos", ["id_recurso", "tipo_recurso", "nome_recurso", "id_categoria_recurso", "desc_categoria_rec", "qtd_estoque", "custo_unitario", "tipo_medida", "data_validade", "data_inclusao", "att_estoque_pedido"], [id_recurso, row.get("TIPO_RECURSO"), row.get("NOME_RECURSO"), as_int(row.get("ID_CATEGORIA_RECURSO")), row.get("DESC_CATEGORIA_REC"), as_decimal(row.get("QTD_ESTOQUE")), as_decimal(row.get("CUSTO_UNITARIO")), row.get("TIPO_MEDIDA"), excel_datetime(row.get("DATA_VALIDADE")), excel_datetime(row.get("DATA_INCLUSAO")), row.get("ATT_ESTOQUE_PEDIDO")], "id_recurso"), "recursos")

    for row in rows(wb["Produtos"]):
        id_produto = as_int(row.get("ID_PRODUTO"))
        if id_produto is not None:
            add(insert("produtos", ["id_produto", "nome_produto", "categoria", "desc_categoria", "preco", "disponibilidade", "peso", "tipo_medida", "data_inclusao", "id_recurso", "rendimento_pessoas"], [id_produto, row.get("NOME_PRODUTO"), as_int(row.get("CATEGORIA")), row.get("DESC_CATEGORIA"), as_decimal(row.get("PRECO")), row.get("DISPONIBILIDADE"), as_decimal(row.get("PESO"), None), row.get("TIPO_MEDIDA"), excel_datetime(row.get("DATA_INCLUSAO")), as_int(row.get("ID_RECURSO")), estimate_yield(row.get("NOME_PRODUTO"), row.get("DESC_CATEGORIA"), row.get("PESO"))], "id_produto"), "produtos")

    for row in rows(wb["Clientes"]):
        id_cliente = as_int(row.get("ID_CLIENTE"))
        if id_cliente is not None:
            add(insert("clientes", ["id_cliente", "nome_cliente", "email_cliente", "telefone_cliente", "endereco_cliente", "id_regiao", "regiao", "data_nascimento", "preferencias_alimentares", "data_inclusao", "entrega"], [id_cliente, row.get("NOME_CLIENTE"), row.get("EMAIL_CLIENTE"), as_phone(row.get("TELEFONE_CLIENTE")), row.get("ENDERECO_CLIENTE"), as_int(row.get("ID_REGIAO")), row.get("REGIAO"), excel_datetime(row.get("DATA_NASCIMENTO")), row.get("PREFERENCIAS_ALIMENTARES"), excel_datetime(row.get("DATA_INCLUSAO")), row.get("ENTREGA")], "id_cliente"), "clientes")

    for row in rows(wb["Pedidos"]):
        id_pedido = as_int(row.get("Ff"))
        if id_pedido is not None:
            add(insert("pedidos", ["id_pedido", "data_pedido", "id_cliente", "nome_cliente", "valor_total", "status_pedido", "id_forma_pagamento", "forma_pagamento", "taxa_cartao", "taxa_entrega", "taxa_embalagem", "taxa_ifood", "valor_final"], [id_pedido, excel_datetime(row.get("DATA_PEDIDO")), as_int(row.get("ID_CLIENTE")), row.get("NOME_CLIENTE"), as_decimal(row.get("VALOR_TOTAL")), status_pedido(row.get("STATUS_PEDIDO")), as_int(row.get("ID_FORMA_PAGAMENTO")), row.get("FORMA_PAGAMENTO"), as_decimal(row.get("TAXA_CARTÃO")), as_decimal(row.get("TAXA_ENTREGA")), as_decimal(row.get("TAXA_EMBALAGEM")), as_decimal(row.get("TAXA_IFOOD")), as_decimal(row.get("VALOR_FINAL"))], "id_pedido"), "pedidos")

    recipe_rows = rows(wb["Receitas"])
    recipe_order = {}
    for row in recipe_rows:
        product_id = as_int(row.get("ID_PRODUTO"))
        if product_id is None:
            continue
        recipe_order.setdefault(product_id, []).append(row)

    order_by_line = {}
    for product_id, product_rows in recipe_order.items():
        sorted_rows = sorted(product_rows, key=lambda item: (classify_order(item.get("NOME_RECURSO")), item["_linha_excel"]))
        for index, item in enumerate(sorted_rows, start=1):
            order_by_line[item["_linha_excel"]] = index

    for row in recipe_rows:
        origem_hash = stable_hash([row.get("ID_RECEITA"), row.get("ID_PRODUTO"), row.get("ID_RECURSO"), row.get("NOME_RECURSO"), row.get("QTD_INGREDIENTE"), row.get("TIPO_MEDIDA")])
        add(insert("receitas", ["id_receita", "id_produto", "nome_produto", "id_recurso", "nome_recurso", "qtd_ingrediente", "tipo_medida", "ordem_preparo", "origem_hash"], [as_int(row.get("ID_RECEITA")), as_int(row.get("ID_PRODUTO")), row.get("NOME_PRODUTO"), as_int(row.get("ID_RECURSO")), row.get("NOME_RECURSO"), as_decimal(row.get("QTD_INGREDIENTE"), None), row.get("TIPO_MEDIDA"), order_by_line.get(row["_linha_excel"]), origem_hash], "origem_hash"), "receitas")

    for row in rows(wb["Itens Pedido"]):
        origem_hash = stable_hash([row.get("ID_PEDIDO"), row.get("ID_PRODUTO"), row.get("NOME_PRODUTO"), row.get("QUANTIDADE"), row.get("PRECO_UNITARIO"), row.get("PRECO_TOTAL"), row.get("OBSERVACAO")])
        add(insert("itens_pedido", ["id_pedido", "id_produto", "nome_produto", "quantidade", "preco_unitario", "preco_total", "observacao", "origem_hash"], [as_int(row.get("ID_PEDIDO")), as_int(row.get("ID_PRODUTO")), row.get("NOME_PRODUTO"), as_decimal(row.get("QUANTIDADE"), None), as_decimal(row.get("PRECO_UNITARIO"), None), as_decimal(row.get("PRECO_TOTAL"), None), row.get("OBSERVACAO"), origem_hash], "origem_hash"), "itens_pedido")

    for row in rows(wb["Compras"]):
        origem_hash = stable_hash([row.get("ID_COMPRA"), row.get("ID_RECURSO"), row.get("NOME_RECURSO"), row.get("ID_FORNECEDOR"), row.get("QTD_COMPRADA"), row.get("DATA_COMPRA"), row.get("CUSTO_TOTAL"), row.get("TIPO_MEDIDA")])
        add(insert("compras", ["id_compra", "id_recurso", "nome_recurso", "id_fornecedor", "nome_fornecedor", "qtd_comprada", "data_compra", "custo_total", "tipo_medida", "origem_hash"], [as_int(row.get("ID_COMPRA")), as_int(row.get("ID_RECURSO")), row.get("NOME_RECURSO"), as_int(row.get("ID_FORNECEDOR")), row.get("NOME_FORNECEDOR"), as_decimal(row.get("QTD_COMPRADA"), None), excel_datetime(row.get("DATA_COMPRA")), as_decimal(row.get("CUSTO_TOTAL"), None), row.get("TIPO_MEDIDA"), origem_hash], "origem_hash"), "compras")

    statements.append("commit;")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"SQL local gerado em {output_path}")
    for table, count in sorted(counts.items()):
        print(f"{table}: {count}")


if __name__ == "__main__":
    main()
