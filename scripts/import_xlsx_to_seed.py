from __future__ import annotations

import datetime as dt
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl nao esta disponivel neste Python.") from exc


DEFAULT_XLSX = Path(r"C:\Users\micha\Downloads\BD_VERSAO_VEGANA Backup 2026-05-29.xlsx")
OUTPUT = Path(__file__).resolve().parents[1] / "supabase" / "seed" / "initial_data.sql"
EXCEL_EPOCH = dt.datetime(1899, 12, 30)


def as_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def as_decimal(value: object, default: str | None = "0") -> str | None:
    if value in (None, ""):
        return default
    try:
        return str(Decimal(str(value)).quantize(Decimal("0.001")).normalize())
    except (InvalidOperation, ValueError):
        return default


def as_phone(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d+(\.\d+)?E\+?\d+", text, flags=re.I):
        try:
            text = str(int(float(text)))
        except ValueError:
            pass
    digits = re.sub(r"\D", "", text)
    return digits or text


def fix_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    try:
        value = value.encode("latin1").decode("utf-8")
    except UnicodeError:
        pass
    return (
        value
        .replace("ÇÇÂO", "CÇÃO")
        .replace("ÇÂO", "ÇÃO")
        .replace("È", "É")
        .replace("è", "é")
    )


def excel_datetime(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.replace(tzinfo=None).isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    if serial <= 0 or serial > 100000:
        return None
    return (EXCEL_EPOCH + dt.timedelta(days=serial)).isoformat()


def sql(value: object) -> str:
    value = fix_text(value)
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value).strip()
    if text == "":
        return "null"
    return "'" + text.replace("'", "''") + "'"


def status_pedido(value: object) -> str:
    text = str(value or "").strip().lower()
    if text == "finalizado":
        return "finalizado"
    if text == "cancelado":
        return "cancelado"
    if text == "rascunho":
        return "rascunho"
    return "pendente"


def rows(ws):
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(item).strip() if item is not None else "" for item in values[0]]
    output = []
    for raw in values[1:]:
        record = {headers[index]: raw[index] if index < len(raw) else None for index in range(len(headers)) if headers[index]}
        if any(value not in (None, "") for value in record.values()):
            output.append(record)
    return output


def insert(table: str, columns: list[str], values: list[object], conflict: str | None = None) -> str:
    rendered_values = ", ".join(sql(value) for value in values)
    rendered_columns = ", ".join(columns)
    statement = f"insert into public.{table} ({rendered_columns}) values ({rendered_values})"
    if conflict:
        update_columns = [column for column in columns if column != conflict]
        updates = ", ".join(f"{column} = excluded.{column}" for column in update_columns)
        statement += f" on conflict ({conflict}) do update set {updates}"
    return statement + ";"


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    statements: list[str] = [
        "-- Seed gerado automaticamente a partir do backup XLSX do Versao Vegana.",
        "-- Tabelas e colunas em portugues, minusculas e sem acentuacao.",
        "begin;",
    ]

    for row in rows(wb["Categorias"]):
        cod_categoria = as_int(row.get("COD_CATEGORIA"))
        if cod_categoria is None:
            continue
        statements.append(insert("categorias", ["cod_categoria", "categoria"], [cod_categoria, row.get("CATEGORIA")], "cod_categoria"))

    for row in rows(wb["Tipo Medida"]):
        tipo_medida = row.get("TIPO_MEDIDA")
        if tipo_medida:
            statements.append(insert("tipos_medida", ["tipo_medida"], [tipo_medida], "tipo_medida"))

    for row in rows(wb["Regiões"]):
        id_regiao = as_int(row.get("ID_REGIAO"))
        if id_regiao is None:
            continue
        statements.append(insert("regioes", ["id_regiao", "regiao", "taxa"], [id_regiao, row.get("REGIAO"), as_decimal(row.get("TAXA"))], "id_regiao"))

    for row in rows(wb["Forma Pagamento"]):
        id_forma_pagamento = as_int(row.get("ID_FORMA_PAGAMENTO"))
        if id_forma_pagamento is None:
            continue
        statements.append(
            insert(
                "formas_pagamento",
                ["id_forma_pagamento", "forma_pagamento", "bandeira", "prazo_recebimento", "taxa"],
                [id_forma_pagamento, row.get("FORMA_PAGAMENTO"), row.get("BANDEIRA"), row.get("PRAZO_RECEBIMENTO"), as_decimal(row.get("TAXA"))],
                "id_forma_pagamento",
            )
        )

    for row in rows(wb["Fornecedores"]):
        id_fornecedor = as_int(row.get("ID_FORNECEDOR"))
        if id_fornecedor is None:
            continue
        statements.append(
            insert(
                "fornecedores",
                ["id_fornecedor", "nome_fornecedor", "telefone_fornecedor", "email_fornecedor", "observacao", "data_inclusao"],
                [id_fornecedor, row.get("NOME_FORNECEDOR"), as_phone(row.get("TELEFONE_FORNECEDOR")), row.get("EMAIL_FORNECEDOR"), row.get("OBSERVACAO"), excel_datetime(row.get("DATA_INCLUSAO"))],
                "id_fornecedor",
            )
        )

    for row in rows(wb["Recursos"]):
        id_recurso = as_int(row.get("ID_RECURSO"))
        if id_recurso is None:
            continue
        statements.append(
            insert(
                "recursos",
                [
                    "id_recurso",
                    "tipo_recurso",
                    "nome_recurso",
                    "id_categoria_recurso",
                    "desc_categoria_rec",
                    "qtd_estoque",
                    "custo_unitario",
                    "tipo_medida",
                    "data_validade",
                    "data_inclusao",
                    "att_estoque_pedido",
                ],
                [
                    id_recurso,
                    row.get("TIPO_RECURSO"),
                    row.get("NOME_RECURSO"),
                    as_int(row.get("ID_CATEGORIA_RECURSO")),
                    row.get("DESC_CATEGORIA_REC"),
                    as_decimal(row.get("QTD_ESTOQUE")),
                    as_decimal(row.get("CUSTO_UNITARIO")),
                    row.get("TIPO_MEDIDA"),
                    excel_datetime(row.get("DATA_VALIDADE")),
                    excel_datetime(row.get("DATA_INCLUSAO")),
                    row.get("ATT_ESTOQUE_PEDIDO"),
                ],
                "id_recurso",
            )
        )

    for row in rows(wb["Produtos"]):
        id_produto = as_int(row.get("ID_PRODUTO"))
        if id_produto is None:
            continue
        statements.append(
            insert(
                "produtos",
                [
                    "id_produto",
                    "nome_produto",
                    "categoria",
                    "desc_categoria",
                    "preco",
                    "disponibilidade",
                    "peso",
                    "tipo_medida",
                    "data_inclusao",
                    "id_recurso",
                ],
                [
                    id_produto,
                    row.get("NOME_PRODUTO"),
                    as_int(row.get("CATEGORIA")),
                    row.get("DESC_CATEGORIA"),
                    as_decimal(row.get("PRECO")),
                    row.get("DISPONIBILIDADE"),
                    as_decimal(row.get("PESO"), None),
                    row.get("TIPO_MEDIDA"),
                    excel_datetime(row.get("DATA_INCLUSAO")),
                    as_int(row.get("ID_RECURSO")),
                ],
                "id_produto",
            )
        )

    for row in rows(wb["Clientes"]):
        id_cliente = as_int(row.get("ID_CLIENTE"))
        if id_cliente is None:
            continue
        statements.append(
            insert(
                "clientes",
                [
                    "id_cliente",
                    "nome_cliente",
                    "email_cliente",
                    "telefone_cliente",
                    "endereco_cliente",
                    "id_regiao",
                    "regiao",
                    "data_nascimento",
                    "preferencias_alimentares",
                    "data_inclusao",
                    "entrega",
                ],
                [
                    id_cliente,
                    row.get("NOME_CLIENTE"),
                    row.get("EMAIL_CLIENTE"),
                    as_phone(row.get("TELEFONE_CLIENTE")),
                    row.get("ENDERECO_CLIENTE"),
                    as_int(row.get("ID_REGIAO")),
                    row.get("REGIAO"),
                    excel_datetime(row.get("DATA_NASCIMENTO")),
                    row.get("PREFERENCIAS_ALIMENTARES"),
                    excel_datetime(row.get("DATA_INCLUSAO")),
                    row.get("ENTREGA"),
                ],
                "id_cliente",
            )
        )

    for row in rows(wb["Pedidos"]):
        id_pedido = as_int(row.get("Ff"))
        if id_pedido is None:
            continue
        statements.append(
            insert(
                "pedidos",
                [
                    "id_pedido",
                    "data_pedido",
                    "id_cliente",
                    "nome_cliente",
                    "valor_total",
                    "status_pedido",
                    "id_forma_pagamento",
                    "forma_pagamento",
                    "taxa_cartao",
                    "taxa_entrega",
                    "taxa_embalagem",
                    "taxa_ifood",
                    "valor_final",
                ],
                [
                    id_pedido,
                    excel_datetime(row.get("DATA_PEDIDO")),
                    as_int(row.get("ID_CLIENTE")),
                    row.get("NOME_CLIENTE"),
                    as_decimal(row.get("VALOR_TOTAL")),
                    status_pedido(row.get("STATUS_PEDIDO")),
                    as_int(row.get("ID_FORMA_PAGAMENTO")),
                    row.get("FORMA_PAGAMENTO"),
                    as_decimal(row.get("TAXA_CARTÃO")),
                    as_decimal(row.get("TAXA_ENTREGA")),
                    as_decimal(row.get("TAXA_EMBALAGEM")),
                    as_decimal(row.get("TAXA_IFOOD")),
                    as_decimal(row.get("VALOR_FINAL")),
                ],
            )
        )

    for row in rows(wb["Itens Pedido"]):
        statements.append(
            insert(
                "itens_pedido",
                ["id_pedido", "id_produto", "nome_produto", "quantidade", "preco_unitario", "preco_total", "observacao"],
                [
                    as_int(row.get("ID_PEDIDO")),
                    as_int(row.get("ID_PRODUTO")),
                    row.get("NOME_PRODUTO"),
                    as_decimal(row.get("QUANTIDADE"), None),
                    as_decimal(row.get("PRECO_UNITARIO"), None),
                    as_decimal(row.get("PRECO_TOTAL"), None),
                    row.get("OBSERVACAO"),
                ],
            )
        )

    for row in rows(wb["Receitas"]):
        statements.append(
            insert(
                "receitas",
                ["id_receita", "id_produto", "nome_produto", "id_recurso", "nome_recurso", "qtd_ingrediente", "tipo_medida"],
                [
                    as_int(row.get("ID_RECEITA")),
                    as_int(row.get("ID_PRODUTO")),
                    row.get("NOME_PRODUTO"),
                    as_int(row.get("ID_RECURSO")),
                    row.get("NOME_RECURSO"),
                    as_decimal(row.get("QTD_INGREDIENTE"), None),
                    row.get("TIPO_MEDIDA"),
                ],
            )
        )

    for row in rows(wb["Produção"]):
        id_producao = as_int(row.get("ID_PRODUCAO"))
        if id_producao is None:
            continue
        statements.append(
            insert(
                "producoes",
                ["id_producao", "id_produto", "nome_produto", "qtd_produzida", "data_producao", "data_validade"],
                [
                    id_producao,
                    as_int(row.get("ID_PRODUTO")),
                    row.get("NOME_PRODUTO"),
                    as_decimal(row.get("QTD_PRODUZIDA"), None),
                    excel_datetime(row.get("DATA_PRODUCAO")),
                    excel_datetime(row.get("DATA_VALIDADE")),
                ],
                "id_producao",
            )
        )

    for row in rows(wb["Compras"]):
        statements.append(
            insert(
                "compras",
                ["id_compra", "id_recurso", "nome_recurso", "id_fornecedor", "nome_fornecedor", "qtd_comprada", "data_compra", "custo_total", "tipo_medida"],
                [
                    as_int(row.get("ID_COMPRA")),
                    as_int(row.get("ID_RECURSO")),
                    row.get("NOME_RECURSO"),
                    as_int(row.get("ID_FORNECEDOR")),
                    row.get("NOME_FORNECEDOR"),
                    as_decimal(row.get("QTD_COMPRADA"), None),
                    excel_datetime(row.get("DATA_COMPRA")),
                    as_decimal(row.get("CUSTO_TOTAL"), None),
                    row.get("TIPO_MEDIDA"),
                ],
            )
        )

    for row in rows(wb["Funcionários"]):
        id_funcionario = as_int(row.get("ID_FUNCIONARIO"))
        if id_funcionario is None:
            continue
        statements.append(
            insert(
                "funcionarios",
                [
                    "id_funcionario",
                    "nome_funcionario",
                    "cargo",
                    "data_admissao",
                    "salario",
                    "cpf",
                    "data_nascimento",
                    "telefone",
                    "email",
                    "endereco",
                    "numero_carteira_trabalho",
                    "status",
                    "observacoes",
                ],
                [
                    id_funcionario,
                    row.get("NOME_FUNCIONARIO"),
                    row.get("CARGO"),
                    excel_datetime(row.get("DATA_ADMISSAO")),
                    as_decimal(row.get("SALARIO"), None),
                    row.get("CPF"),
                    excel_datetime(row.get("DATA_NASCIMENTO")),
                    as_phone(row.get("TELEFONE")),
                    row.get("EMAIL"),
                    row.get("ENDERECO"),
                    row.get("NUMERO_CARTEIRA_TRABALHO"),
                    row.get("STATUS"),
                    row.get("OBSERVACOES"),
                ],
                "id_funcionario",
            )
        )

    statements.append("commit;")
    OUTPUT.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"Seed gerado em {OUTPUT}")
    print(f"{len(statements) - 4} inserts criados.")


if __name__ == "__main__":
    main()
